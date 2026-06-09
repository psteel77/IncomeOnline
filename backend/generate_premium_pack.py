"""
Premium Pack bundler — $12.99 product.

Bundle = the 10 free MoneyRules guides (print-ready PDF) PLUS 10 premium items:
  • 6 INTERACTIVE Excel calculators with live, auto-updating charts
       (type your numbers in the highlighted cells → the graph redraws)
  • 4 premium-only PDF guides
  • a welcome / how-to-use letter (PDF)

Excel calculators are generated fresh at runtime (openpyxl works in production).
The PDF guides + welcome letter are pre-built by build_guide_pdfs.py (needs
LibreOffice + Montserrat) and committed to /static/premium, so build_premium_pack
is runtime-safe.
"""
import io
import os
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference

from moneyrules_template import (
    create_moneyrules_document, add_title_page, add_styled_heading,
    add_body_text, add_highlight_box, add_branded_table, add_closing_page,
    save_to_buffer, BODY_TEXT,
)
from docx.shared import Pt

STATIC_DIR = os.path.join(os.path.dirname(__file__), 'static')
PREMIUM_DIR = os.path.join(STATIC_DIR, 'premium')
PREMIUM_FILE = 'MoneyRules_Premium_Pack.zip'
PREMIUM_PATH = os.path.join(STATIC_DIR, PREMIUM_FILE)

# Style palette
PURPLE_HEX = '7C3AED'
PINK_HEX = 'DB2777'
AMBER_HEX = 'F59E0B'
GREEN_HEX = '059669'
DARK_HEX = '1F2937'
INPUT_HEX = 'FEF3C7'   # light amber — marks "type here" cells

_THIN = Side(border_style='thin', color='CCCCCC')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ---------------------------------------------------------------
# Shared spreadsheet helpers
# ---------------------------------------------------------------

def _header_style(cell, fill_hex=PURPLE_HEX):
    cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=fill_hex)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _BORDER


def _title_row(ws, row, text, col_span, fill_hex=PURPLE_HEX):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name='Georgia', size=16, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=fill_hex)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _input_cell(cell, fmt='£#,##0'):
    """Highlight a user-input cell so it's obvious where to type."""
    cell.fill = PatternFill('solid', fgColor=INPUT_HEX)
    cell.font = Font(bold=True, color='92400E')
    cell.border = _BORDER
    cell.number_format = fmt


def _label(cell):
    cell.font = Font(bold=True, color='374151')


def _line_chart(ws, title, data_ref, cats_ref, anchor, x_title='Year', y_title='£'):
    ch = LineChart()
    ch.title = title
    ch.style = 12
    ch.height = 8.5
    ch.width = 17
    ch.x_axis.title = x_title
    ch.y_axis.title = y_title
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    ws.add_chart(ch, anchor)


def _bar_chart(ws, title, data_ref, cats_ref, anchor, x_title='', y_title='£'):
    ch = BarChart()
    ch.type = 'col'
    ch.title = title
    ch.style = 10
    ch.height = 8.5
    ch.width = 17
    ch.x_axis.title = x_title
    ch.y_axis.title = y_title
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    ws.add_chart(ch, anchor)


def _pie_chart(ws, title, data_ref, cats_ref, anchor):
    ch = PieChart()
    ch.title = title
    ch.height = 8.5
    ch.width = 11
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    ws.add_chart(ch, anchor)


def _save(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ===============================================================
# INTERACTIVE CALCULATOR 1 — Investment Growth & Compound Interest
# ===============================================================

def _calc_investment_growth():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Investment Growth'
    _title_row(ws, 1, 'MoneyRules · Investment Growth Calculator', 4)

    _label(ws.cell(row=3, column=1, value='Starting amount £'))
    _label(ws.cell(row=4, column=1, value='Monthly contribution £'))
    _label(ws.cell(row=5, column=1, value='Expected annual return %'))
    ws['B3'] = 10000; _input_cell(ws['B3'])
    ws['B4'] = 200;   _input_cell(ws['B4'])
    ws['B5'] = 7;     _input_cell(ws['B5'], '0.0')
    ws.cell(row=6, column=1, value='Type your own numbers in the amber cells').font = Font(italic=True, size=9, color='92400E')

    # Projection table (0..30 years)
    hdr = ['Year', 'Total Deposited £', 'Interest Earned £', 'Balance £']
    for i, h in enumerate(hdr, 1):
        _header_style(ws.cell(row=8, column=i))
        ws.cell(row=8, column=i, value=h)
    ws['A9'] = 0
    ws['B9'] = '=$B$3'
    ws['C9'] = 0
    ws['D9'] = '=$B$3'
    for r in range(10, 40):
        ws.cell(row=r, column=1, value=f'=A{r-1}+1')
        ws.cell(row=r, column=2, value=f'=B{r-1}+$B$4*12')
        ws.cell(row=r, column=4, value=f'=D{r-1}*(1+$B$5/100)+$B$4*12*(1+$B$5/200)')
        ws.cell(row=r, column=3, value=f'=D{r}-B{r}')
        for c in (2, 3, 4):
            ws.cell(row=r, column=c).number_format = '£#,##0'

    data = Reference(ws, min_col=2, max_col=4, min_row=8, max_row=39)
    cats = Reference(ws, min_col=1, min_row=9, max_row=39)
    _line_chart(ws, 'Your money over time', data, cats, 'F3')

    _autosize(ws, [26, 18, 18, 16])
    return _save(wb)


# ===============================================================
# INTERACTIVE CALCULATOR 2 — Budget Command Center (50/30/20)
# ===============================================================

def _calc_budget_command_center():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Budget'
    _title_row(ws, 1, 'MoneyRules · Budget Command Center (50/30/20)', 3)

    _label(ws.cell(row=3, column=1, value='Monthly take-home pay £'))
    ws['B3'] = 2500; _input_cell(ws['B3'])
    ws.cell(row=3, column=3, value='type your pay here').font = Font(italic=True, size=9, color='92400E')

    sections = [
        ('NEEDS  (target 50%)', PURPLE_HEX, ['Rent / Mortgage', 'Council tax', 'Utilities', 'Groceries', 'Transport', 'Insurance', 'Min. debt payments']),
        ('WANTS  (target 30%)', PINK_HEX, ['Eating out', 'Subscriptions', 'Shopping', 'Entertainment', 'Travel']),
        ('SAVINGS  (target 20%)', GREEN_HEX, ['Emergency fund', 'ISA / Pension', 'Goal savings']),
    ]
    row = 5
    subtotal_cells = []
    for name, color, items in sections:
        ws.cell(row=row, column=1, value=name).font = Font(bold=True, size=12, color=color)
        row += 1
        start = row
        for it in items:
            ws.cell(row=row, column=1, value=it)
            ws.cell(row=row, column=2, value=0)
            _input_cell(ws.cell(row=row, column=2))
            row += 1
        ws.cell(row=row, column=1, value=f'{name.split("  ")[0]} total').font = Font(italic=True, bold=True)
        ws.cell(row=row, column=2, value=f'=SUM(B{start}:B{row-1})').number_format = '£#,##0'
        subtotal_cells.append(row)
        row += 2

    # Summary table for chart
    s = row + 1
    ws.cell(row=s, column=1, value='Category').font = Font(bold=True)
    ws.cell(row=s, column=2, value='You spend £').font = Font(bold=True)
    ws.cell(row=s, column=3, value='Target £').font = Font(bold=True)
    for i, (lbl, sub_row, pct) in enumerate(zip(['Needs', 'Wants', 'Savings'], subtotal_cells, [0.5, 0.3, 0.2])):
        rr = s + 1 + i
        ws.cell(row=rr, column=1, value=lbl)
        ws.cell(row=rr, column=2, value=f'=B{sub_row}').number_format = '£#,##0'
        ws.cell(row=rr, column=3, value=f'=$B$3*{pct}').number_format = '£#,##0'

    data = Reference(ws, min_col=2, max_col=3, min_row=s, max_row=s + 3)
    cats = Reference(ws, min_col=1, min_row=s + 1, max_row=s + 3)
    _bar_chart(ws, 'You spend vs your 50/30/20 target', data, cats, 'E5', y_title='£/month')
    pie_data = Reference(ws, min_col=2, min_row=s, max_row=s + 3)
    _pie_chart(ws, 'Where your money goes', pie_data, cats, 'E22')

    _autosize(ws, [26, 16, 16])
    return _save(wb)


# ===============================================================
# INTERACTIVE CALCULATOR 3 — Debt Payoff Planner (Snowball/Avalanche)
# ===============================================================

def _calc_debt_payoff():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Debt Payoff'
    _title_row(ws, 1, 'MoneyRules · Debt Payoff Planner', 6, fill_hex=PINK_HEX)

    hdr = ['Debt Name', 'Balance £', 'Min £/mo', 'APR %', 'Extra £/mo', 'Months to clear']
    for i, h in enumerate(hdr, 1):
        _header_style(ws.cell(row=3, column=i), PINK_HEX)
        ws.cell(row=3, column=i, value=h)
    samples = [
        ('Store card', 380, 25, 29.9, 250),
        ('Credit card A', 1200, 40, 19.9, 0),
        ('Overdraft', 1500, 30, 35.0, 0),
        ('Personal loan', 5400, 180, 12.5, 0),
        ('Car finance', 3200, 150, 9.9, 0),
    ]
    first, last = 4, 4 + len(samples) - 1
    for i, (n, bal, mn, apr, extra) in enumerate(samples, start=4):
        ws.cell(row=i, column=1, value=n); _input_cell(ws.cell(row=i, column=1), 'General')
        ws.cell(row=i, column=2, value=bal); _input_cell(ws.cell(row=i, column=2))
        ws.cell(row=i, column=3, value=mn); _input_cell(ws.cell(row=i, column=3))
        ws.cell(row=i, column=4, value=apr); _input_cell(ws.cell(row=i, column=4), '0.0')
        ws.cell(row=i, column=5, value=extra); _input_cell(ws.cell(row=i, column=5))
        ws.cell(row=i, column=6, value=f'=ROUNDUP(B{i}/(C{i}+E{i}),0)')

    trow = last + 1
    ws.cell(row=trow, column=1, value='TOTAL DEBT').font = Font(bold=True)
    ws.cell(row=trow, column=2, value=f'=SUM(B{first}:B{last})').number_format = '£#,##0'
    ws.cell(row=trow, column=2).font = Font(bold=True, color=PINK_HEX)

    note1 = 'SNOWBALL: pay minimums on all, throw every spare £ at the SMALLEST balance first (fast wins, momentum).'
    note2 = 'AVALANCHE: throw every spare £ at the HIGHEST APR first (saves the most interest). Sort rows accordingly.'
    for k, note in enumerate((note1, note2)):
        rr = trow + 2 + k
        ws.cell(row=rr, column=1, value=note).font = Font(italic=True, size=10, color='6B7280')
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)

    data = Reference(ws, min_col=2, min_row=3, max_row=last)
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    _bar_chart(ws, 'Balance by debt (target smallest or highest-APR first)', data, cats, 'H3', y_title='£')

    _autosize(ws, [22, 16, 14, 10, 14, 16])
    return _save(wb)


# ===============================================================
# INTERACTIVE CALCULATOR 4 — Net Worth Tracker (12-month trend)
# ===============================================================

def _calc_net_worth_tracker():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Net Worth'
    _title_row(ws, 1, 'MoneyRules · Net Worth Tracker', 4)

    ws.cell(row=3, column=1, value='Update your totals once a month — watch the trend line climb.').font = Font(italic=True, size=10, color='6B7280')
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)

    hdr = ['Month', 'Total Assets £', 'Total Liabilities £', 'Net Worth £']
    for i, h in enumerate(hdr, 1):
        _header_style(ws.cell(row=5, column=i))
        ws.cell(row=5, column=i, value=h)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    first = 6
    for i, m in enumerate(months):
        r = first + i
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=0); _input_cell(ws.cell(row=r, column=2))
        ws.cell(row=r, column=3, value=0); _input_cell(ws.cell(row=r, column=3))
        ws.cell(row=r, column=4, value=f'=B{r}-C{r}').number_format = '£#,##0'
    last = first + len(months) - 1

    data = Reference(ws, min_col=4, min_row=5, max_row=last)
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    _line_chart(ws, 'Your net worth trend', data, cats, 'F5', x_title='Month')

    _autosize(ws, [12, 18, 20, 16])
    return _save(wb)


# ===============================================================
# INTERACTIVE CALCULATOR 5 — Retirement / FIRE Calculator
# ===============================================================

def _calc_fire():
    wb = Workbook()
    ws = wb.active
    ws.title = 'FIRE Calculator'
    _title_row(ws, 1, 'MoneyRules · Retirement / FIRE Calculator', 4)

    labels = [
        ('Annual spending in retirement £', 30000, '£#,##0'),
        ('Current amount invested £', 20000, '£#,##0'),
        ('Monthly contribution £', 500, '£#,##0'),
        ('Expected annual return %', 7, '0.0'),
        ('Safe withdrawal rate %', 4, '0.0'),
    ]
    for i, (lbl, val, fmt) in enumerate(labels, start=3):
        _label(ws.cell(row=i, column=1, value=lbl))
        ws.cell(row=i, column=2, value=val)
        _input_cell(ws.cell(row=i, column=2), fmt)
    ws.cell(row=9, column=1, value='YOUR FIRE NUMBER').font = Font(bold=True, color=PURPLE_HEX, size=12)
    ws.cell(row=9, column=2, value='=B3/(B7/100)').number_format = '£#,##0'
    ws.cell(row=9, column=2).font = Font(bold=True, color=PURPLE_HEX, size=12)
    ws.cell(row=10, column=1, value='(= annual spend / withdrawal rate, e.g. 25x at 4%)').font = Font(italic=True, size=9, color='6B7280')

    hdr = ['Year', 'Portfolio £', 'FIRE Target £']
    for i, h in enumerate(hdr, 1):
        _header_style(ws.cell(row=12, column=i))
        ws.cell(row=12, column=i, value=h)
    ws['A13'] = 0
    ws['B13'] = '=$B$4'
    ws['C13'] = '=$B$3/($B$7/100)'
    for r in range(14, 54):  # 40 years
        ws.cell(row=r, column=1, value=f'=A{r-1}+1')
        ws.cell(row=r, column=2, value=f'=B{r-1}*(1+$B$6/100)+$B$5*12*(1+$B$6/200)')
        ws.cell(row=r, column=3, value='=$B$3/($B$7/100)')
        ws.cell(row=r, column=2).number_format = '£#,##0'
        ws.cell(row=r, column=3).number_format = '£#,##0'

    data = Reference(ws, min_col=2, max_col=3, min_row=12, max_row=53)
    cats = Reference(ws, min_col=1, min_row=13, max_row=53)
    _line_chart(ws, 'When your portfolio crosses your FIRE target = freedom', data, cats, 'E3')

    _autosize(ws, [30, 16, 16])
    return _save(wb)


# ===============================================================
# INTERACTIVE CALCULATOR 6 — Mortgage & Loan Repayment Calculator
# ===============================================================

def _calc_mortgage():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mortgage Calc'
    _title_row(ws, 1, 'MoneyRules · Mortgage & Loan Calculator', 4)

    _label(ws.cell(row=3, column=1, value='Loan / mortgage amount £'))
    _label(ws.cell(row=4, column=1, value='Annual interest rate %'))
    _label(ws.cell(row=5, column=1, value='Term (years)'))
    ws['B3'] = 200000; _input_cell(ws['B3'])
    ws['B4'] = 5.0;    _input_cell(ws['B4'], '0.00')
    ws['B5'] = 25;     _input_cell(ws['B5'], '0')

    ws.cell(row=7, column=1, value='Monthly payment').font = Font(bold=True, color=PURPLE_HEX)
    ws['B7'] = '=PMT(B4/100/12,B5*12,-B3)'
    ws['B7'].number_format = '£#,##0.00'
    ws['B7'].font = Font(bold=True, color=PURPLE_HEX, size=12)
    ws.cell(row=8, column=1, value='Total interest over term').font = Font(bold=True)
    ws['B8'] = '=B7*B5*12-B3'
    ws['B8'].number_format = '£#,##0'

    hdr = ['Year', 'Remaining Balance £']
    for i, h in enumerate(hdr, 1):
        _header_style(ws.cell(row=10, column=i))
        ws.cell(row=10, column=i, value=h)
    ws['A11'] = 0
    ws['B11'] = '=$B$3'
    for r in range(12, 47):  # up to 35 years; floors at 0
        yr = r - 11
        ws.cell(row=r, column=1, value=yr)
        ws.cell(row=r, column=2, value=(
            f'=MAX(0,$B$3*(1+$B$4/100/12)^({yr}*12)'
            f'-$B$7*(((1+$B$4/100/12)^({yr}*12)-1)/($B$4/100/12)))'
        )).number_format = '£#,##0'

    data = Reference(ws, min_col=2, min_row=10, max_row=46)
    cats = Reference(ws, min_col=1, min_row=11, max_row=46)
    _line_chart(ws, 'Your balance falling to zero', data, cats, 'D3')

    _autosize(ws, [26, 20])
    return _save(wb)


CALCULATORS = [
    ('01_Investment_Growth_Calculator.xlsx', _calc_investment_growth),
    ('02_Budget_Command_Center.xlsx',        _calc_budget_command_center),
    ('03_Debt_Payoff_Planner.xlsx',          _calc_debt_payoff),
    ('04_Net_Worth_Tracker.xlsx',            _calc_net_worth_tracker),
    ('05_Retirement_FIRE_Calculator.xlsx',   _calc_fire),
    ('06_Mortgage_Loan_Calculator.xlsx',     _calc_mortgage),
]


# ---------------------------------------------------------------
# PDF guide content helpers
# ---------------------------------------------------------------

def _bullets(doc, items):
    for it in items:
        p = doc.add_paragraph(it, style='List Bullet')
        for r in p.runs:
            r.font.size = Pt(11)
            r.font.color.rgb = BODY_TEXT


def _steps(doc, items):
    for it in items:
        p = doc.add_paragraph(it, style='List Number')
        for r in p.runs:
            r.font.size = Pt(11)
            r.font.color.rgb = BODY_TEXT


# ===============================================================
# PREMIUM PDF GUIDE 1 — The Investor's Starter Kit
# ===============================================================

def _guide_investors_starter_kit():
    doc = create_moneyrules_document(title="The Investor's Starter Kit", subtitle='Index Funds, ETFs & Shares — Explained Simply')
    add_title_page(doc, title="Investor's Starter Kit",
        subtitle='Index Funds, ETFs & Shares Explained',
        tagline='Everything a beginner needs to start investing\nwith confidence — without the jargon.')

    add_styled_heading(doc, 'Why Invest At All?', level=1)
    add_body_text(doc, 'Cash in a savings account quietly loses value to inflation every year. Investing puts your money to work in real businesses so it can grow faster than prices rise. Over decades, that difference is life-changing — but only if you start, stay invested, and keep costs low.')
    add_highlight_box(doc, 'Time in the market beats timing the market.\nStarting early matters more than starting big.')
    doc.add_page_break()

    add_styled_heading(doc, 'The Building Blocks', level=1)
    add_branded_table(doc,
        headers=['Type', 'What it is', 'Best for'],
        data=[
            ('Index fund', 'Owns every company in an index (e.g. FTSE All-World)', 'Hands-off, long-term core'),
            ('ETF', 'An index fund that trades like a share', 'Low fees, flexibility'),
            ('Individual share', 'A stake in one company', 'Higher risk, optional slice'),
            ('Bond fund', 'Lends to governments/companies for interest', 'Stability nearer retirement'),
        ])
    add_body_text(doc, 'For most people, a single low-cost global index fund or ETF is 90% of the answer. It instantly spreads your money across thousands of companies in dozens of countries.')
    doc.add_page_break()

    add_styled_heading(doc, 'The Three Rules That Matter Most', level=1)
    _steps(doc, [
        'Keep costs low — an ongoing fee of 0.2% vs 1.0% can cost you tens of thousands over a lifetime.',
        'Diversify — never bet your future on one company or one country. A global fund does this automatically.',
        'Stay invested — markets fall sometimes. The investors who win are the ones who do not panic-sell.',
    ])
    add_highlight_box(doc, 'A 1% annual fee can quietly eat ~25% of your final pot over 30 years. Costs are the one thing you can control.')
    doc.add_page_break()

    add_styled_heading(doc, 'How To Actually Start (This Week)', level=1)
    _steps(doc, [
        'Open a Stocks & Shares ISA with a low-cost platform.',
        'Choose ONE diversified global index fund or ETF to begin.',
        'Set up a monthly direct debit — even £25 builds the habit.',
        'Turn on automatic reinvestment of dividends.',
        'Ignore the daily noise. Review once or twice a year, not daily.',
    ])
    add_body_text(doc, 'Use the Investment Growth Calculator in this pack to see how your monthly contribution could grow over 10, 20 and 30 years.')
    add_closing_page(doc)
    return save_to_buffer(doc)


# ===============================================================
# PREMIUM PDF GUIDE 2 — The Buy-to-Let & Property Profit Guide
# ===============================================================

def _guide_buy_to_let():
    doc = create_moneyrules_document(title='Buy-to-Let & Property Profit', subtitle='How UK Rental Property Really Makes Money')
    add_title_page(doc, title='Buy-to-Let Profit Guide',
        subtitle='How UK Rental Property Really Makes Money',
        tagline='Yields, mortgages, tax and the numbers\nthat decide whether a deal is worth it.')

    add_styled_heading(doc, 'The Two Ways Property Pays', level=1)
    add_body_text(doc, 'Rental property earns in two ways: monthly rental income (cash flow) and long-term growth in the property value (capital appreciation). A good deal stacks up on cash flow today, with growth as the bonus — never the other way around.')
    add_highlight_box(doc, 'Buy on the numbers, not on the dream.\nIf it does not cash-flow, it is a liability, not an asset.')
    doc.add_page_break()

    add_styled_heading(doc, 'Know Your Yields', level=1)
    add_branded_table(doc,
        headers=['Metric', 'Formula', 'Healthy range'],
        data=[
            ('Gross yield', 'Annual rent / property price', '5%-8%+'),
            ('Net yield', '(Rent - costs) / price', '3%-6%+'),
            ('ROI', 'Annual profit / cash invested', '8%-15%+'),
        ])
    add_body_text(doc, 'Gross yield is the headline; net yield is the truth. Always work out net yield after mortgage interest, insurance, management, maintenance and void periods.')
    doc.add_page_break()

    add_styled_heading(doc, 'The Costs Beginners Forget', level=1)
    _bullets(doc, [
        'Stamp Duty — including the surcharge on additional properties.',
        'Mortgage arrangement & broker fees.',
        'Letting agent fees (typically 8-12% of rent if fully managed).',
        'Maintenance & a sinking fund for big repairs (boiler, roof).',
        'Void periods — budget for roughly one month empty per year.',
        'Landlord insurance, gas/electrical safety certificates, EPC.',
    ])
    doc.add_page_break()

    add_styled_heading(doc, 'The Tax Reality (Section 24)', level=1)
    add_body_text(doc, 'Since the "Section 24" changes, individual landlords can no longer deduct mortgage interest from rental income before tax — instead you get a 20% tax credit. For higher-rate taxpayers this is a big deal, and it is why some landlords hold property through a limited company. Always take professional advice before buying.')
    add_highlight_box(doc, 'Run every deal through the numbers AFTER tax.\nA good gross yield can be a loss-maker after Section 24.')
    doc.add_page_break()

    add_styled_heading(doc, 'Your First-Deal Checklist', level=1)
    _steps(doc, [
        'Set your strategy: cash flow, growth, or both.',
        'Pick an area with strong rental demand and realistic prices.',
        'Calculate gross AND net yield before viewing.',
        'Get a mortgage agreement in principle.',
        'Stress-test the deal: +2% interest, 2 months void — does it still work?',
        'Budget all the costs above, then make your offer on the numbers.',
    ])
    add_closing_page(doc)
    return save_to_buffer(doc)


# ===============================================================
# PREMIUM PDF GUIDE 3 — The Tax-Efficiency Masterclass
# ===============================================================

def _guide_tax_efficiency():
    doc = create_moneyrules_document(title='Tax-Efficiency Masterclass', subtitle='Keep More of What You Earn — Legally')
    add_title_page(doc, title='Tax-Efficiency Masterclass',
        subtitle='Keep More of What You Earn — Legally',
        tagline='The allowances, wrappers and simple moves\nthat cut your tax bill without any risk.')

    add_styled_heading(doc, 'The Principle', level=1)
    add_body_text(doc, 'Tax efficiency is not about dodging tax — it is about using the allowances and accounts the government deliberately created to reward saving and investing. Most people leave hundreds or thousands on the table every year simply by not using them.')
    add_highlight_box(doc, 'It is not what you earn — it is what you keep.\nEvery allowance you waste is gone for good at year end.')
    doc.add_page_break()

    add_styled_heading(doc, 'Your Key Annual Allowances', level=1)
    add_body_text(doc, 'Allowances reset each tax year and most cannot be carried forward. Treat them like a use-it-or-lose-it budget. (Figures change each year — always check the current limits.)')
    add_branded_table(doc,
        headers=['Allowance', 'What it does'],
        data=[
            ('Personal Allowance', 'Income you can earn before any income tax'),
            ('ISA Allowance', 'Save/invest each year completely tax-free'),
            ('Pension Annual Allowance', 'Contributions that get tax relief'),
            ('Dividend Allowance', 'Dividends before dividend tax applies'),
            ('Capital Gains Allowance', 'Profit on investments before CGT'),
            ('Personal Savings Allowance', 'Bank interest before tax'),
        ])
    doc.add_page_break()

    add_styled_heading(doc, 'The Two Wrappers Everyone Should Use', level=1)
    add_styled_heading(doc, 'The ISA', level=2)
    add_body_text(doc, 'Everything inside an ISA grows free of income tax and capital gains tax, and withdrawals are tax-free. Flexible, simple, and ideal for medium-term goals and tax-free income later.')
    add_styled_heading(doc, 'The Pension (SIPP / workplace)', level=2)
    add_body_text(doc, 'Contributions get tax relief at your marginal rate — a basic-rate taxpayer turns £80 into £100 instantly; a higher-rate taxpayer can reclaim more. Locked away until pension age, but the relief is the single biggest legal boost to your wealth.')
    doc.add_page_break()

    add_styled_heading(doc, 'Quick Wins Most People Miss', level=1)
    _bullets(doc, [
        'Claim higher-rate pension tax relief through self-assessment — it is not always automatic.',
        'Marriage Allowance — transfer unused personal allowance between spouses.',
        'Salary sacrifice for pension — saves income tax AND National Insurance.',
        "Use both partners' ISA and CGT allowances to double your tax-free room.",
        'Hold investments inside an ISA/pension rather than a taxable account.',
    ])
    add_highlight_box(doc, 'A single afternoon setting these up can be worth more than a pay rise — and it repeats every single year.')
    add_body_text(doc, 'This guide is educational, not personal tax advice. For your own situation, check current HMRC limits or speak to a qualified adviser.')
    add_closing_page(doc)
    return save_to_buffer(doc)


# ===============================================================
# PREMIUM PDF GUIDE 4 — The 12-Month Money Makeover
# ===============================================================

def _guide_money_makeover():
    doc = create_moneyrules_document(title='The 12-Month Money Makeover', subtitle='One Focused Step a Month to Transform Your Finances')
    add_title_page(doc, title='12-Month Money Makeover',
        subtitle='One Step a Month to a Total Transformation',
        tagline='A simple, do-able plan — one focus per month.\nFollow it and your finances will look different in a year.')

    add_styled_heading(doc, 'How To Use This Plan', level=1)
    add_body_text(doc, 'Big money goals fail because we try to fix everything at once. This plan gives you ONE focus each month. Do that month task, then relax — you are allowed to ignore everything else. Twelve small wins compound into a complete transformation.')
    add_highlight_box(doc, 'You do not need willpower — you need a sequence.\nOne month, one move.')
    doc.add_page_break()

    add_styled_heading(doc, 'Your 12-Month Map', level=1)
    add_branded_table(doc,
        headers=['Month', 'Focus'],
        data=[
            ('1', 'Track every penny — know exactly where your money goes'),
            ('2', 'Build your 50/30/20 budget (use the Budget tool)'),
            ('3', 'Start a 1,000 starter emergency fund'),
            ('4', 'List all debts; choose snowball or avalanche'),
            ('5', 'Cancel/renegotiate bills, subscriptions & insurance'),
            ('6', 'Open & fund a Stocks & Shares ISA'),
            ('7', 'Boost your pension contribution by 1-3%'),
            ('8', 'Grow emergency fund to 3 months of essentials'),
            ('9', 'Claim every allowance (ISA, marriage, pension relief)'),
            ('10', 'Start a side-hustle or ask for a pay rise'),
            ('11', 'Write a will & set pension nominations'),
            ('12', 'Review net worth; set next year goals'),
        ])
    doc.add_page_break()

    add_styled_heading(doc, 'The First Quarter In Detail', level=1)
    add_styled_heading(doc, 'Month 1 — See The Truth', level=2)
    add_body_text(doc, 'Record every single transaction for 30 days. No judgement, no changes yet — just awareness. You cannot fix what you cannot see.')
    add_styled_heading(doc, 'Month 2 — Give Every Pound a Job', level=2)
    add_body_text(doc, 'Split your take-home pay into 50% needs, 30% wants, 20% savings. The Budget Command Center spreadsheet in this pack does the maths for you.')
    add_styled_heading(doc, 'Month 3 — Build a Buffer', level=2)
    add_body_text(doc, 'Save a 1,000 starter emergency fund as fast as you can. This one buffer stops a flat tyre or broken boiler from becoming new debt.')
    doc.add_page_break()

    add_styled_heading(doc, 'Make It Stick', level=1)
    _bullets(doc, [
        'Automate everything — pay yourself first on payday.',
        "Put this month's focus somewhere you see it daily.",
        'Track your net worth monthly (use the Net Worth Tracker).',
        "Celebrate each month's win — momentum is the secret ingredient.",
    ])
    add_highlight_box(doc, 'In 12 months: a budget, an emergency fund, a debt plan, investments started, and a will. That is a different life.')
    add_closing_page(doc)
    return save_to_buffer(doc)


# Premium PDF guides: output filename in /static/premium -> generator fn
PREMIUM_GUIDES = [
    ('A1_Investors_Starter_Kit.pdf',      _guide_investors_starter_kit),
    ('A2_Buy_to_Let_Property_Profit.pdf', _guide_buy_to_let),
    ('A3_Tax_Efficiency_Masterclass.pdf', _guide_tax_efficiency),
    ('A4_12_Month_Money_Makeover.pdf',    _guide_money_makeover),
]


# ===============================================================
# README / welcome letter
# ===============================================================

def _readme_docx():
    doc = create_moneyrules_document(title='MoneyRules Premium Pack', subtitle='Welcome & Contents')
    add_title_page(doc, title='Welcome',
        subtitle='MoneyRules Premium Pack',
        tagline='Thank you for supporting Income Online!\nHere is everything inside and how to use it.')

    add_styled_heading(doc, "What's Inside", level=1)
    add_body_text(doc, 'Your Premium Pack contains the 10 free MoneyRules guides PLUS 10 premium items: 6 interactive Excel calculators with live charts, and 4 premium-only PDF guides. The calculators are fully editable — type your own numbers into the highlighted cells and the graphs redraw instantly.')

    add_styled_heading(doc, 'Interactive Calculators (Excel — type & watch the chart move)', level=2)
    add_branded_table(doc,
        headers=['#', 'Tool', 'What its chart shows'],
        data=[
            ('1', 'Investment Growth Calculator', 'Your pot growing year by year'),
            ('2', 'Budget Command Center (50/30/20)', 'Where your money goes vs target'),
            ('3', 'Debt Payoff Planner', 'Balance per debt (snowball/avalanche)'),
            ('4', 'Net Worth Tracker', '12-month net worth trend line'),
            ('5', 'Retirement / FIRE Calculator', 'When you cross your freedom number'),
            ('6', 'Mortgage & Loan Calculator', 'Your balance falling to zero'),
        ])

    add_styled_heading(doc, 'Premium Guides (PDF)', level=2)
    add_branded_table(doc,
        headers=['#', 'Guide'],
        data=[
            ('7', "The Investor's Starter Kit"),
            ('8', 'The Buy-to-Let & Property Profit Guide'),
            ('9', 'The Tax-Efficiency Masterclass'),
            ('10', 'The 12-Month Money Makeover'),
        ])

    add_styled_heading(doc, 'How to Use', level=1)
    add_body_text(doc, 'Open the Excel calculators in Microsoft Excel, Google Sheets, or Apple Numbers. Type your own figures into the amber/highlighted cells — every formula and chart updates automatically. The PDF guides are print-ready; read one, act on its checklist, then move to the next.')
    add_highlight_box(doc, 'Pair each guide with its tool:\nInvesting guide + Growth Calculator. Makeover + Budget & Net Worth trackers.')

    add_styled_heading(doc, 'Stay in Touch', level=1)
    add_body_text(doc, 'Visit www.incomeonline.info to discover 199+ platforms for earning online. New guides ship regularly — as a Premium member you will hear about them first.')
    add_closing_page(doc)
    return save_to_buffer(doc)


# ===============================================================
# Bundle everything into a single ZIP
# ===============================================================

def build_premium_pack():
    """
    Assemble the Premium Pack ZIP into /static.

    PDF guides (10 free + 4 premium + welcome) are pre-built & committed in
    /static and /static/premium by build_guide_pdfs.py. The 6 Excel calculators
    are generated fresh here (openpyxl is runtime-safe). Raises loudly if any
    required PDF is missing so we never ship an incomplete pack.
    """
    os.makedirs(STATIC_DIR, exist_ok=True)

    free_guides = [
        ('01_Rule_of_72.pdf',              'The_Rule_of_72_Guide.pdf'),
        ('02_50-30-20_Budget.pdf',         'The_50_30_20_Budget_Rule.pdf'),
        ('03_Passive_Income.pdf',          'Passive_Income_Beginners_Guide.pdf'),
        ('04_Debt_Snowball.pdf',           'The_Debt_Snowball_Method.pdf'),
        ('05_Emergency_Fund.pdf',          'The_Emergency_Fund_Guide.pdf'),
        ('06_Compound_Interest.pdf',       'Compound_Interest_Handbook.pdf'),
        ('07_UK_Tax_Basics.pdf',           'UK_Tax_Basics_Freelancers.pdf'),
        ('08_UK_Credit_Score.pdf',         'UK_Credit_Score_Masterclass.pdf'),
        ('09_ISA_vs_SIPP.pdf',             'ISA_vs_SIPP_Complete_Guide.pdf'),
        ('10_Side_Hustle_Quick_Start.pdf', 'Side_Hustle_Quick_Start_Guide.pdf'),
    ]

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        welcome = os.path.join(PREMIUM_DIR, '00_WELCOME_START_HERE.pdf')
        if not os.path.exists(welcome):
            raise FileNotFoundError(f"Missing {welcome}. Run `python build_guide_pdfs.py` first.")
        with open(welcome, 'rb') as f:
            zf.writestr('00_WELCOME_START_HERE.pdf', f.read())

        for archive_name, static_file in free_guides:
            src = os.path.join(STATIC_DIR, static_file)
            if not os.path.exists(src):
                raise FileNotFoundError(f"Missing {src}. Run `python build_guide_pdfs.py` first.")
            with open(src, 'rb') as f:
                zf.writestr(f'10 Free Guides/{archive_name}', f.read())

        for out_name, _fn in PREMIUM_GUIDES:
            src = os.path.join(PREMIUM_DIR, out_name)
            if not os.path.exists(src):
                raise FileNotFoundError(f"Missing {src}. Run `python build_guide_pdfs.py` first.")
            with open(src, 'rb') as f:
                zf.writestr(f'Premium Guides/{out_name}', f.read())

        for out_name, gen_fn in CALCULATORS:
            zf.writestr(f'Interactive Tools/{out_name}', gen_fn().read())

    buf.seek(0)
    with open(PREMIUM_PATH, 'wb') as f:
        f.write(buf.read())

    size_mb = os.path.getsize(PREMIUM_PATH) / 1024 / 1024
    print(f"Premium Pack built: {PREMIUM_PATH} ({size_mb:.2f} MB)")
    return PREMIUM_PATH


if __name__ == '__main__':
    build_premium_pack()
